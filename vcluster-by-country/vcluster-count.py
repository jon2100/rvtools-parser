#!/usr/bin/env python3

import os
import pandas as pd
import argparse
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# Conversion factors
MIB_TO_MB = 1.048576
MB_TO_GB = 1024
MB_TO_TB = 1048576

# Helper function to find column with multiple possible names
def find_column(df, possible_names):
    for name in possible_names:
        if name in df.columns:
            return name
    return None

# Helper function to find a sheet with multiple possible names
def find_sheet(xl, possible_sheet_names):
    for sheet_name in possible_sheet_names:
        if sheet_name in xl.sheet_names:
            return sheet_name
    return None

# Function to read the mapping file (Country, vCenter, vCluster)
def read_mapping(file_path, sheet_name="vClusterLoc"):
    try:
        xl = pd.ExcelFile(file_path)
        sheet_name = find_sheet(xl, [sheet_name])
        if not sheet_name:
            print(f"Sheet '{sheet_name}' not found in {file_path}.")
            return pd.DataFrame()
        
        df = xl.parse(sheet_name)
        if 'Country' not in df.columns or 'vCenter' not in df.columns or 'vCluster' not in df.columns:
            print(f"Required columns ('Country', 'vCenter', 'vCluster') not found in {file_path}.")
            return pd.DataFrame()
        
        # Clean up vCluster names by trimming whitespace
        df['vCluster'] = df['vCluster'].str.strip().str.lower()
        df = df.sort_values(by=['Country', 'vCluster']).reset_index(drop=True)
        return df
    except Exception as e:
        print(f"Error reading mapping file {file_path}: {e}")
        return pd.DataFrame()

# Function to get VM counts and hardware totals (CPUs, Memory, Disk) from the "vInfo" worksheet and standalone VMs
def count_vms_in_info(file_path, possible_sheet_names=["vInfo"]):
    try:
        xl = pd.ExcelFile(file_path)
        sheet_name = find_sheet(xl, possible_sheet_names)
        if not sheet_name:
            print(f"None of the sheets {possible_sheet_names} found in {file_path}.")
            return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()

        df = xl.parse(sheet_name)
        print(f"Columns in '{sheet_name}' of '{file_path}': {df.columns.tolist()}")  # Debugging: Show columns

        # Find necessary columns
        cluster_col = find_column(df, ['Cluster'])
        vm_col = find_column(df, ['VM'])
        vcenter_col = find_column(df, ['VI SDK Server'])
        cpu_col = find_column(df, ['CPUs'])
        memory_col = find_column(df, ['Memory'])
        disk_col = find_column(df, ['Total disk capacity MiB'])
        os_col = find_column(df, ['OS according to the configuration file'])
        os_tools_col = find_column(df, ['OS according to the VMware Tools'])

        if not all([cluster_col, vm_col, vcenter_col, cpu_col, memory_col, disk_col, os_col, os_tools_col]):
            print(f"Required columns not found in '{file_path}'.")
            return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()

        # Apply filtering logic
        if 'Template' in df.columns and 'SRM Placeholder' in df.columns:
            df = df[(df['Template'] != True) & (df['SRM Placeholder'] != True)]

        df = df[~df[os_col].astype(str).str.contains('Template|SRM Placeholder', case=False, na=False)]
        df = df[df[os_tools_col] != "VMware Photon OS (64-bit)"]

        # Extract vCenter name
        df['vCenter'] = df[vcenter_col].apply(lambda x: x.split('.')[0] if pd.notna(x) else 'Unknown')
        df['vCluster'] = df[cluster_col].fillna('None - StandAlone').str.strip().str.lower()

        # Aggregate VM counts by vCluster
        vm_counts = df.groupby(['vCenter', 'vCluster']).size().reset_index(name='VM_Count')

        # Aggregate hardware totals
        hardware_totals = df.groupby(['vCenter', 'vCluster']).agg(
            Total_CPUs=(cpu_col, 'sum'),
            Total_Memory_GB=(memory_col, lambda x: round(x.sum() / MB_TO_GB, 2)),
            Total_Disk_TB=(disk_col, lambda x: round((x.sum() * MIB_TO_MB) / MB_TO_TB, 2))
        ).reset_index()

        # Extract standalone VMs
        standalone_vms = df[df['vCluster'] == 'none - standalone'][['vCenter', 'VM']]

        return vm_counts[['vCenter', 'vCluster', 'VM_Count']], standalone_vms, hardware_totals
    except Exception as e:
        print(f"Error processing 'vInfo' worksheet in '{file_path}': {e}")
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()

# Function to get host counts from the "vCluster" worksheet
def count_hosts_in_vcluster(file_path, possible_sheet_names=["vCluster"]):
    try:
        xl = pd.ExcelFile(file_path)
        sheet_name = find_sheet(xl, possible_sheet_names)
        if not sheet_name:
            print(f"None of the sheets {possible_sheet_names} found in {file_path}.")
            return pd.DataFrame()

        df = xl.parse(sheet_name)
        print(f"Columns in '{sheet_name}' of '{file_path}': {df.columns.tolist()}")

        # Find necessary columns
        cluster_col = find_column(df, ['Name'])
        numhosts_col = find_column(df, ['NumHosts'])

        if not cluster_col or not numhosts_col:
            print(f"Required columns ('Name', 'NumHosts') not found in '{file_path}'.")
            return pd.DataFrame()

        # Clean up cluster names
        df['vCluster'] = df[cluster_col].str.strip().str.lower()

        # Aggregate host counts by vCluster
        host_counts = df.groupby('vCluster').agg(Host_Count=(numhosts_col, 'sum')).reset_index()

        return host_counts
    except Exception as e:
        print(f"Error processing 'vCluster' worksheet in '{file_path}': {e}")
        return pd.DataFrame()

# Function to add a new row with the correct number of columns
def add_row_to_df(output_df, values, columns):
    row_data = values + [''] * (len(columns) - len(values))
    return pd.DataFrame([row_data], columns=columns)

# Main function to process data and combine results
def process_data(src_dir, mapping_file, mapping_sheet, output_file):
    combined_df = pd.DataFrame()
    standalone_vms_df = pd.DataFrame()

    # Read the mapping file if provided
    mapping_df = pd.DataFrame()
    use_country = False
    if mapping_file:
        mapping_df = read_mapping(mapping_file, sheet_name=mapping_sheet)
        if not mapping_df.empty:
            use_country = True  # Set flag to use country column if the mapping file is valid

    # Loop through each file in the source directory
    for file_name in os.listdir(src_dir):
        if file_name.endswith(".xlsx"):
            file_path = os.path.join(src_dir, file_name)
            print(f"Processing file: {file_path}")

            # Get VM counts and hardware totals from "vInfo" worksheet
            vm_counts, standalone_vms, hardware_totals = count_vms_in_info(file_path)
            
            # Get host counts from "vCluster" worksheet
            host_counts = count_hosts_in_vcluster(file_path)
            
            # Check if VM counts DataFrame has valid data
            if not vm_counts.empty:
                cluster_data = pd.merge(vm_counts, host_counts, on='vCluster', how='left').fillna(0)
                cluster_data = pd.merge(cluster_data, hardware_totals, on=['vCenter', 'vCluster'], how='left').fillna(0)
                combined_df = pd.concat([combined_df, cluster_data], ignore_index=True)
                standalone_vms_df = pd.concat([standalone_vms_df, standalone_vms], ignore_index=True)
            else:
                print(f"No valid VM data found in '{file_path}'.")

    if combined_df.empty:
        print("No valid cluster data found to process. Exiting.")
        return

    # Merge with mapping data if it exists
    if use_country:
        final_df = pd.merge(mapping_df, combined_df, on=['vCenter', 'vCluster'], how='right')
        final_df['Country'] = final_df['Country'].fillna('Unknown')
    else:
        final_df = combined_df
        final_df['Country'] = 'Unknown'  # Default value when mapping is not provided

    final_df['Host_Count'] = final_df['Host_Count'].fillna(0).astype(int)
    final_df['VM_Count'] = final_df['VM_Count'].fillna(0).astype(int)

    # Add the "Site" column based on the vCluster naming convention
    final_df['Site'] = final_df['vCluster'].apply(lambda x: 'EDC' if 'dc1h1' in x or 'dc2h2' in x else 'Plan')

    # Add the "Group" column based on new grouping criteria
    def assign_group(vcluster):
        if '-infra-dr' in vcluster:
            return 'Infra-DR Group'
        elif vcluster.startswith('dc1h') or vcluster.startswith('dc2h'):
            if '-edge-infra' in vcluster:  # Include '-edge-infra' clusters
                return vcluster[:4]  # Group by 'dc1h' or 'dc2h'
        return vcluster[:4]  # Default grouping by the first 4 characters
    
    final_df['Group'] = final_df['vCluster'].apply(assign_group)

    # Group by relevant columns for aggregation
    group_by_columns = ['Country', 'vCenter', 'Group', 'vCluster', 'Site'] if use_country else ['vCenter', 'Group', 'vCluster', 'Site']
    grouped_df = final_df.groupby(group_by_columns).agg(
        Total_Hosts=('Host_Count', 'sum'),
        Total_VMs=('VM_Count', 'sum'),
        Total_CPUs=('Total_CPUs', 'sum'),
        Total_Memory_GB=('Total_Memory_GB', 'sum'),
        Total_Disk_TB=('Total_Disk_TB', 'sum')
    ).reset_index()

    # Format columns
    for col in ['Total_Hosts', 'Total_VMs', 'Total_CPUs']:
        grouped_df[col] = grouped_df[col].apply(lambda x: f"{int(x):,}")
    for col in ['Total_Memory_GB', 'Total_Disk_TB']:
        grouped_df[col] = grouped_df[col].apply(lambda x: f"{x:,.2f}")

    # Prepare final output DataFrame with totals
    output_df = pd.DataFrame(columns=grouped_df.columns)
    last_country = None
    last_group = None
    country_host_total = country_vm_total = country_cpu_total = country_memory_total = country_disk_total = 0
    group_host_total = group_vm_total = group_cpu_total = group_memory_total = group_disk_total = 0

    # Loop through grouped_df to generate the output DataFrame with totals and spacing
    for _, row in grouped_df.iterrows():
        # Add row separation and group totals logic
        if use_country and last_country and last_country != row['Country']:
            # Add group totals if any non-zero values exist
            if last_group is not None and (group_host_total != 0 or group_vm_total != 0 or group_cpu_total != 0 or group_memory_total != 0.0 or group_disk_total != 0.0):
                output_df = pd.concat([output_df, add_row_to_df(output_df, ['Group Totals:', f"{group_host_total:,}", f"{group_vm_total:,}", f"{group_cpu_total:,}", f"{group_memory_total:,.2f}", f"{group_disk_total:,.2f}"], output_df.columns)], ignore_index=True)

            # Add country totals
            if country_host_total != 0 or country_vm_total != 0 or country_cpu_total != 0 or country_memory_total != 0.0 or country_disk_total != 0.0:
                output_df = pd.concat([output_df, add_row_to_df(output_df, ['Country Totals:', f"{country_host_total:,}", f"{country_vm_total:,}", f"{country_cpu_total:,}", f"{country_memory_total:,.2f}", f"{country_disk_total:,.2f}"], output_df.columns)], ignore_index=True)

            # Add spacing rows
            output_df = pd.concat([output_df, pd.DataFrame([[''] * len(output_df.columns)], columns=output_df.columns)], ignore_index=True)
            output_df = pd.concat([output_df, pd.DataFrame([[''] * len(output_df.columns)], columns=output_df.columns)], ignore_index=True)

            country_host_total = country_vm_total = country_cpu_total = country_memory_total = country_disk_total = 0

        # Add group totals when the group changes
        if last_group and last_group != row['Group']:
            # Only add group totals if they are non-zero
            if group_host_total != 0 or group_vm_total != 0 or group_cpu_total != 0 or group_memory_total != 0.0 or group_disk_total != 0.0:
                output_df = pd.concat([output_df, add_row_to_df(output_df, ['Group Totals:', f"{group_host_total:,}", f"{group_vm_total:,}", f"{group_cpu_total:,}", f"{group_memory_total:,.2f}", f"{group_disk_total:,.2f}"], output_df.columns)], ignore_index=True)

            # Add spacing rows (always add these rows even if no country)
            output_df = pd.concat([output_df, pd.DataFrame([[''] * len(output_df.columns)], columns=output_df.columns)], ignore_index=True)
            output_df = pd.concat([output_df, pd.DataFrame([[''] * len(output_df.columns)], columns=output_df.columns)], ignore_index=True)

            group_host_total = group_vm_total = group_cpu_total = group_memory_total = group_disk_total = 0

        # Add the current row
        output_df = pd.concat([output_df, pd.DataFrame([row], columns=grouped_df.columns)], ignore_index=True)

        # Update totals
        if use_country:
            country_host_total += int(row['Total_Hosts'].replace(',', ''))
            country_vm_total += int(row['Total_VMs'].replace(',', ''))
            country_cpu_total += int(row['Total_CPUs'].replace(',', ''))
            country_memory_total += float(row['Total_Memory_GB'].replace(',', ''))
            country_disk_total += float(row['Total_Disk_TB'].replace(',', ''))

        group_host_total += int(row['Total_Hosts'].replace(',', ''))
        group_vm_total += int(row['Total_VMs'].replace(',', ''))
        group_cpu_total += int(row['Total_CPUs'].replace(',', ''))
        group_memory_total += float(row['Total_Memory_GB'].replace(',', ''))
        group_disk_total += float(row['Total_Disk_TB'].replace(',', ''))

        last_country = row['Country'] if use_country else None
        last_group = row['Group']

    # Add last group and country totals
    if last_group is not None and (group_host_total != 0 or group_vm_total != 0 or group_cpu_total != 0 or group_memory_total != 0.0 or group_disk_total != 0.0):
        output_df = pd.concat([output_df, add_row_to_df(output_df, ['Group Totals:', f"{group_host_total:,}", f"{group_vm_total:,}", f"{group_cpu_total:,}", f"{group_memory_total:,.2f}", f"{group_disk_total:,.2f}"], output_df.columns)], ignore_index=True)

    if use_country and (country_host_total != 0 or country_vm_total != 0 or country_cpu_total != 0 or country_memory_total != 0.0 or country_disk_total != 0.0):
        output_df = pd.concat([output_df, add_row_to_df(output_df, ['Country Totals:', f"{country_host_total:,}", f"{country_vm_total:,}", f"{country_cpu_total:,}", f"{country_memory_total:,.2f}", f"{country_disk_total:,.2f}"], output_df.columns)], ignore_index=True)

    # Save to Excel
    if not output_file.endswith(".xlsx"):
        output_file += ".xlsx"
    
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        output_df.to_excel(writer, sheet_name="vCluster Summary", index=False)
        standalone_vms_df.to_excel(writer, sheet_name="Standalone_VMs", index=False)

    # Adjust column widths and alignment
    adjust_column_widths_and_alignment(output_file)
    print(f"Results saved to {output_file}")

# Function to adjust column widths and alignment
def adjust_column_widths_and_alignment(output_file):
    try:
        workbook = load_workbook(output_file)
        for sheet in workbook.sheetnames:
            worksheet = workbook[sheet]

            for col in worksheet.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        # Calculate max length considering header as well
                        max_length = max(max_length, len(str(cell.value)))
                        
                        # Set cell alignment
                        if column in ['D', 'E', 'F', 'G', 'H']:  # Center alignment for numeric columns
                            cell.alignment = Alignment(horizontal='center')
                        elif column == 'C':  # Left alignment for vCluster
                            cell.alignment = Alignment(horizontal='left')
                    except:
                        pass
                adjusted_width = (max_length + 2)  # Add some padding to the width
                worksheet.column_dimensions[column].width = adjusted_width

        workbook.save(output_file)
    except Exception as e:
        print(f"Error adjusting column widths and alignment: {e}")

# Main function
def main():
    parser = argparse.ArgumentParser(description='Process Excel files and summarize vClusters by Country and vCenter.')
    parser.add_argument('-s', '--src', type=str, default='./data', help='Source directory containing vCluster Excel files.')
    parser.add_argument('-m', '--mapping', type=str, help='Path to the Excel file containing Country, vCenter, and vCluster mapping (optional).')
    parser.add_argument('-ms', '--mapping-sheet', type=str, default='vClusterLoc', help='Sheet name of the mapping data in the Excel file.')
    parser.add_argument('-d', '--dst', type=str, default='./output', help='Destination directory for output file.')
    parser.add_argument('-n', '--name', type=str, default='output.xlsx', help='Output file name.')

    args = parser.parse_args()

    # Ensure the output directory exists
    os.makedirs(args.dst, exist_ok=True)
    
    # Define the output file path
    output_file = os.path.join(args.dst, args.name)

    # Process the source directory and generate output
    process_data(args.src, args.mapping, args.mapping_sheet, output_file)

if __name__ == "__main__":
    main()
