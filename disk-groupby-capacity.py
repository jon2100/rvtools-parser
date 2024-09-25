#!/usr/bin/env python3

import pandas as pd
import os
import argparse
import openpyxl
from openpyxl.chart import PieChart, Reference
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart.label import DataLabelList
from openpyxl.worksheet.table import Table, TableStyleInfo
from concurrent.futures import ProcessPoolExecutor, as_completed
from tqdm import tqdm
from openpyxl import load_workbook

# Conversion factor for MiB to MB
MIB_TO_MB = 1.048576

def process_file(file_path, os_filters, capacity_ranges):
    """Process a single Excel file and return OS counts for all capacity ranges and VMware Photon OS counts."""
    df = pd.read_excel(file_path)
    
    os_col_candidates = df.columns[df.columns.str.contains("OS according to the configuration file", case=False)]
    if os_col_candidates.empty:
        return None, None, None

    os_col = os_col_candidates.tolist()[0]

    capacity_col_candidates = df.columns[df.columns.str.contains("Total disk capacity MiB|Total disk capacity MB", case=False, regex=True)]
    if capacity_col_candidates.empty:
        return None, None, None

    capacity_col = capacity_col_candidates.tolist()[0]

    if 'Template' in df.columns and 'SRM Placeholder' in df.columns:
        df = df[(df['Template'] != True) & (df['SRM Placeholder'] != True)]

    df = df[~df[os_col].astype(str).str.contains('Template|SRM Placeholder|AdditionalBackEnd', case=False, na=False)]

    if "MiB" in capacity_col:
        df[capacity_col] = df[capacity_col] * MIB_TO_MB

    vmware_os_col_candidates = df.columns[df.columns.str.contains("OS according to the VMware Tools", case=False)]
    photon_result = None
    if not vmware_os_col_candidates.empty:
        vmware_os_col = vmware_os_col_candidates.tolist()[0]
        photon_result = df[df[vmware_os_col] == "VMware Photon OS (64-bit)"]

    results_by_range = {}
    for min_capacity, max_capacity, label in capacity_ranges:
        filtered_df = df[
            (df[capacity_col] >= min_capacity) &
            (df[capacity_col] <= max_capacity)
        ]
        if not filtered_df.empty:
            grouped_result = filtered_df.groupby(os_col).size().reset_index(name='Count')
            grouped_result['Capacity Range'] = label
            grouped_result['OS according to the configuration file'] = grouped_result[os_col]
            results_by_range[label] = grouped_result

    return results_by_range, photon_result, df[os_col].dropna().unique()

def parallel_process_files(file_paths, capacity_ranges):
    """Process files in parallel, returning the combined OS results for all capacity ranges and VMware Photon OS results."""
    all_results_by_range = {label: [] for _, _, label in capacity_ranges}
    photon_results = []
    unique_os_filters = set()

    with ProcessPoolExecutor() as executor:
        future_to_file = {executor.submit(process_file, file_path, [], capacity_ranges): file_path for file_path in file_paths}
        for future in tqdm(as_completed(future_to_file), total=len(future_to_file), desc="Processing files in parallel"):
            try:
                results_by_range, photon_result, unique_os = future.result()
                for label, result in results_by_range.items():
                    if result is not None:
                        all_results_by_range[label].append(result)

                if photon_result is not None and not photon_result.empty:
                    photon_results.append(photon_result)

                if unique_os is not None:
                    unique_os_filters.update(unique_os)
            except Exception as exc:
                print(f"File {future_to_file[future]} generated an exception: {exc}")

    combined_results_by_range = {}
    for label, results in all_results_by_range.items():
        if results:
            combined_df = pd.concat(results, ignore_index=True)
            combined_results_by_range[label] = combined_df.groupby(combined_df.columns[0]).sum().reset_index()
        else:
            combined_results_by_range[label] = pd.DataFrame(columns=["OS according to the configuration file", "Count"])

    if photon_results:
        photon_combined_df = pd.concat(photon_results, ignore_index=True)
        photon_summary = photon_combined_df.groupby("OS according to the VMware Tools").size().reset_index(name='Count')
    else:
        photon_summary = pd.DataFrame(columns=["OS according to the VMware Tools", "Count"])

    return combined_results_by_range, photon_summary, unique_os_filters

def insert_break_and_sum(df):
    """Insert sum row and ensure column exists."""
    total_count = df['Count'].sum()
    break_df = pd.DataFrame({'OS according to the configuration file': ['Disk OS Sum', ''], 'Count': [total_count, '']})
    df = pd.concat([df, break_df], ignore_index=True)
    return df

def create_pie_chart(wb, df, title):
    """Create a pie chart and pivot table in the worksheet."""
    ws = wb.create_sheet(title)
    chart = PieChart()
    chart.title = title

    try:
        # Filter out the rows that don't have a 'Capacity Range'
        df = df[df['Capacity Range'].notna()]

        # Group the data by 'Capacity Range' and calculate the sum of 'Count'
        pivot_table = df.groupby('Capacity Range')['Count'].sum().reset_index()

        # Add a row for the total
        total_count = pivot_table['Count'].sum()
        pivot_table.loc[len(pivot_table)] = ['Total', total_count]

        # Create a pivot table in the worksheet
        ws.append(["Capacity Range", "Count"])  # Updated column headers
        for r in dataframe_to_rows(pivot_table, index=False, header=True):
            ws.append(r)

        # Create a pie chart
        labels = Reference(ws, min_col=1, min_row=2, max_row=len(pivot_table) + 1)  # Updated max row
        data = Reference(ws, min_col=2, min_row=2, max_row=len(pivot_table) + 1)  # Updated max row
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(labels)
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showPercent = True
        ws.add_chart(chart, "E2")
    except KeyError as e:
        print(f"The column {e} does not exist in the DataFrame.")

def set_column_widths(sheet):
    """Set the width of the columns to fit the header length."""
    for col in sheet.columns:
        max_length = 0
        col_letter = col[0].column_letter  # Get the column letter
        for cell in col:
            try:  # Get the max length of the column values and header
                max_length = max(len(str(cell.value)), max_length)
            except:
                pass
        adjusted_width = max_length + 2  # Add some padding for better readability
        sheet.column_dimensions[col_letter].width = adjusted_width

def main():
    parser = argparse.ArgumentParser(description="Process Excel files and generate OS disk capacity reports.")
    
    # Add both short and long arguments with proper help text
    parser.add_argument('-s', '--src', default='./data', help='Source folder containing Excel files (default: ./data)')
    parser.add_argument('-d', '--dst', default='./output', help='Destination folder for the output file (default: ./output)')
    parser.add_argument('-n', '--name', default='output', help="Base name of the output file without extension (default: output)")
    parser.add_argument('-h', '--help', action='help', default=argparse.SUPPRESS, help='Displays this help')

    args = parser.parse_args()

    src_folder = os.path.abspath(args.src)
    dst_folder = os.path.abspath(args.dst)
    output_file = os.path.join(dst_folder, f"{args.name}{file_extension}")
    # Example usage in the terminal:
    # python3 disk-groupby-capacity.py -s /path/to/source -d /path/to/destination -n os_report

    # Get all Excel files in the source directory
    file_paths = [os.path.join(src_folder, f) for f in os.listdir(src_folder) if f.endswith('.xlsx')]

    capacity_ranges = [
        (0, 149, '0 MB - 149 MB'),
        (150, 2000000, '150 MB - 2 TB'),
        (2000001, 10000000, '2 TB - 10 TB'),
        (10000001, 20000000, '10 TB - 20 TB'),
        (20000001, 40000000, '20 TB - 40 TB')
    ]

    os.makedirs(dst_folder, exist_ok=True)

    combined_results_by_range, photon_combined, unique_os_filters = parallel_process_files(file_paths, capacity_ranges)

    if file_extension == '.xlsx':
        wb = Workbook()
        ws = wb.active

        for i, (label, df) in enumerate(combined_results_by_range.items()):
            df = insert_break_and_sum(df)
            create_pie_chart(wb, df, label)

        ws = wb.create_sheet("VMware Photon OS")
        ws.append(["OS according to the VMware Tools", "Count"])
        for r in dataframe_to_rows(photon_combined, index=False, header=True):
            ws.append(r)
        create_pie_chart(wb, photon_combined, "VMware Photon OS")

        wb.save(output_file)
    else:
        with open(output_file, 'w', newline='') as f:
            writer = csv.writer(f)
            for label, df in combined_results_by_range.items():
                df = insert_break_and_sum(df)
                writer.writerow([label])
                writer.writerow(["OS according to the configuration file", "Count"])
                writer.writerows(df.values.tolist())
                writer.writerow([])

<<<<<<< HEAD:disk-groupby-capacity.py
    # Insert sum row for Photon OS if photon_combined is not empty
    if not photon_combined.empty:
        photon_combined['Capacity Range'] = 'All Capacities'
        photon_combined['OS according to the configuration file'] = 'VMware Photon OS (64-bit)'
        photon_count = photon_combined['Count'].sum()
        photon_total_row = pd.DataFrame({
            'OS according to the configuration file': ['Disk OS Sum'],
            'Count': [photon_count],
            'Capacity Range': ['All Capacities']
        })
        combined_results = pd.concat([combined_results, photon_combined, photon_total_row], ignore_index=True)

    # Rearrange columns to switch "OS according to the configuration file" to the first column
    if 'OS according to the configuration file' in combined_results.columns and 'OS according to the VMware Tools' in combined_results.columns:
        columns_order = ['OS according to the configuration file', 'Count', 'Capacity Range', 'OS according to the VMware Tools']
        combined_results = combined_results[columns_order]

    # Output the combined result to a single Excel file using the openpyxl engine
    output_file = os.path.join(dst_folder, f"{args.name}.xlsx")
    combined_results.to_excel(output_file, index=False, engine='openpyxl')

    # Load the file and set column widths
    workbook = load_workbook(output_file)
    sheet = workbook.active
    set_column_widths(sheet)
    workbook.save(output_file)

    print(f"Combined results including VMware Photon OS saved to {output_file}")
=======
            writer.writerow(["VMware Photon OS"])
            writer.writerow(["OS according to the VMware Tools", "Count"])
            writer.writerows(photon_combined.values.tolist())
>>>>>>> main:checkme3.py

if __name__ == "__main__":
    main()